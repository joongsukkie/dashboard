"""
AI-Powered Data Analytics Agent
Flask web application for marketing and digital analytics workflows.
"""
from __future__ import annotations

import os
import io
import json
import uuid
import base64
import logging
import traceback
from datetime import datetime
from threading import Lock

import numpy as np
import pandas as pd
from flask import (
    Flask, render_template, request, jsonify, session, send_file, abort
)
from werkzeug.utils import secure_filename

# Charting
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from plotly.utils import PlotlyJSONEncoder

# Stats
from scipy import stats

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table,
    TableStyle, Image as RLImage
)

# RAG (ReturnLens diagnostic engine — used only when archetype == 'reviews')
import rag

# B2C archetype detection + playbooks (senior-analyst per-dataset-type analysis)
import archetypes as arch_mod
import playbooks as playbook_mod

# Synthetic Research — persona-grounded synthetic market surveys
import personas as personas_mod
import synthetic_research as sr_mod
import evidence as evidence_mod


# -----------------------------------------------------------------------------
# App setup
# -----------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("analytics-agent")

app = Flask(__name__)
# Stable secret key so session cookies survive restarts / across workers.
# If FLASK_SECRET_KEY isn't set, fall back to a random key (dev mode only).
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or os.urandom(32)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# SESSION_COOKIE_SECURE is left at default (False) so local HTTP dev works;
# on HTTPS hosts (Render/Railway) browsers still accept non-Secure cookies.

# In-memory store keyed by session token. Not persisted.
STORE: dict = {}
STORE_LOCK = Lock()

ACCENT = "#15803D"  # Emerald-700
# Single-accent palette: used for single-series charts (time series, scatter).
PALETTE = ["#15803D", "#22C55E", "#14532D", "#4E7C4A", "#65A30D", "#0D9488", "#4B5650", "#121815"]
# Distinct categorical palette: each category gets a visually different hue.
# Used for grouped bars, box plots by category, and binary comparisons where
# the colors must be readable as separate groups (not shades of green).
CAT_PALETTE = [
    "#15803D", "#D97706", "#2563EB", "#DC2626", "#7C3AED",
    "#0891B2", "#DB2777", "#CA8A04", "#4B5650", "#059669",
    "#9333EA", "#EA580C",
]


# -----------------------------------------------------------------------------
# Session helpers
# -----------------------------------------------------------------------------
def get_sid() -> str:
    if "sid" not in session:
        session["sid"] = uuid.uuid4().hex
    return session["sid"]


def get_state() -> dict:
    sid = get_sid()
    with STORE_LOCK:
        if sid not in STORE:
            STORE[sid] = {}
        return STORE[sid]


# -----------------------------------------------------------------------------
# Data cleaning
# -----------------------------------------------------------------------------
ID_NAME_RE = __import__("re").compile(r"(?i)(^|_)id$|_id$|^id_|(order|customer|transaction|txn|user|account|invoice|record)_?(id|no|num|number)?$")


def _detect_id_columns(df: pd.DataFrame) -> list[str]:
    """Identify primary-key-like columns by name and uniqueness ratio."""
    ids = []
    for c in df.columns:
        name_match = bool(ID_NAME_RE.search(str(c)))
        try:
            uniq_ratio = df[c].nunique(dropna=True) / max(1, len(df))
        except TypeError:
            uniq_ratio = 0.0
        if name_match and uniq_ratio > 0.6:
            ids.append(c)
    return ids


def _fuzzy_canonicalize(series: pd.Series, cutoff: float = 0.85) -> tuple[pd.Series, dict]:
    """Merge near-duplicate categorical labels (typos, case, whitespace).

    Canonical form = most frequent original spelling in each similarity
    cluster. Returns the replaced series plus a mapping {variant: canonical}
    for audit.
    """
    import difflib
    s = series.astype(str).str.strip()
    s_norm = s.str.lower()
    counts = s_norm.value_counts()
    labels = counts.index.tolist()

    clusters: list[list[str]] = []
    used: set[str] = set()
    for label in labels:
        if label in used:
            continue
        group = [label]
        used.add(label)
        for other in labels:
            if other in used:
                continue
            if difflib.SequenceMatcher(None, label, other).ratio() >= cutoff:
                group.append(other)
                used.add(other)
        clusters.append(group)

    # Canonical spelling heuristic:
    #   1. Prefer the LONGEST form — typos usually drop letters (hyderbad →
    #      Hyderabad, hyd → hyderabad).
    #   2. Among equal-length candidates, prefer a mixed-case form (has an
    #      uppercase letter) over all-lowercase — proper nouns typically
    #      capitalize.
    #   3. Frequency is only the final tiebreaker.
    mapping_norm: dict[str, str] = {}
    audit: dict[str, str] = {}
    for group in clusters:
        cluster_mask = s_norm.isin(group)
        variant_counts = s[cluster_mask].value_counts()
        def _rank(v: str) -> tuple:
            has_upper = any(ch.isupper() for ch in v)
            return (len(v), 1 if has_upper else 0, int(variant_counts[v]))
        canonical = max(variant_counts.index, key=_rank)
        for norm in group:
            mapping_norm[norm] = canonical
            orig_variants = s[s_norm == norm].unique()
            for v in orig_variants:
                if v != canonical:
                    audit[v] = canonical
    return s_norm.map(mapping_norm), audit


import re as _re

# Patterns used by cleaning + PII detection.
_PII_PATTERNS = {
    "email":  _re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"),
    "phone":  _re.compile(r"(?:\+?\d{1,3}[\s.-]?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}"),
    "card":   _re.compile(r"\b(?:\d[ -]?){13,19}\b"),
    "ssn_us": _re.compile(r"\b\d{3}-\d{2}-\d{4}\b"),
}

# Common analyst placeholders for "missing" — when a number appears far more
# often than any other and looks like one of these, it's almost always a
# sentinel, not a real measurement.
_PLACEHOLDER_NUMS = {-1, -999, -9999, 999, 9999, 99999, 999999}


def _parse_numeric_smart(series: pd.Series) -> tuple[pd.Series, str | None]:
    """Best-effort numeric parsing for currency, EU decimals, accounting
    parentheses-as-negative, trailing percent signs and currency codes.

    Returns (parsed_series, label) where label is one of:
      'numeric', 'numeric (currency)', 'numeric (percent)',
      'numeric (eu)', 'numeric (accounting)', or None if parsing failed.
    """
    s_raw = series.astype(str).str.strip()
    sample = s_raw.dropna().head(300)
    if sample.empty:
        return series, None

    has_percent  = sample.str.contains("%").any()
    has_currency = sample.str.contains(r"[$£€¥₹]|\b(?:USD|EUR|GBP|JPY|INR)\b",
                                       regex=True, case=False).any()
    has_parens   = sample.str.match(r"^\(.*\)$").any()
    has_eu_fmt   = sample.str.match(r"^-?\d{1,3}(\.\d{3})+,\d+$").any()

    s = s_raw
    if has_eu_fmt:
        # 1.234,56 → 1234.56
        s = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    else:
        # Strip thousands separators (commas) only when there's no EU pattern.
        s = s.str.replace(",", "", regex=False)

    if has_parens:
        # (123.45) → -123.45
        s = s.where(~s.str.match(r"^\(.*\)$"),
                    "-" + s.str.replace(r"[()]", "", regex=True))

    # Strip currency symbols, percent signs, trailing currency codes.
    s = s.str.replace(r"[$£€¥₹]", "", regex=True)
    s = s.str.replace(r"(?i)\b(?:USD|EUR|GBP|JPY|INR|CAD|AUD)\b", "", regex=True)
    s = s.str.replace("%", "", regex=False).str.strip()

    parsed = pd.to_numeric(s, errors="coerce")
    coverage = parsed.notna().sum() / max(1, len(s_raw.dropna()))
    if coverage < 0.9:
        return series, None

    if has_percent:
        return parsed, "numeric (percent)"
    if has_eu_fmt:
        return parsed, "numeric (eu)"
    if has_parens:
        return parsed, "numeric (accounting)"
    if has_currency:
        return parsed, "numeric (currency)"
    return parsed, "numeric"


def _parse_dates_smart(series: pd.Series) -> tuple[pd.Series | None, str | None]:
    """Try date parsing with a few strategies. Returns (parsed_or_None, label).

    Strategies (in order):
    1. Excel serial numbers (integers in the 25000–60000 range).
    2. ISO/standard via pandas' default parser.
    3. dayfirst=True (DD/MM/YYYY common outside the US).
    4. Mixed formats: pd.to_datetime(..., format='mixed') for pandas >= 2.0.
    """
    sample = series.dropna().astype(str).head(300)
    if sample.empty:
        return None, None

    # Skip if it looks numeric (so we don't accidentally datetime-ize prices).
    if sample.str.match(r"^-?\d+(\.\d+)?$").mean() > 0.8:
        # Could still be Excel serials — check range.
        try:
            as_num = pd.to_numeric(sample, errors="coerce")
            if as_num.between(25000, 60000).mean() > 0.9:
                origin = pd.Timestamp("1899-12-30")
                parsed = origin + pd.to_timedelta(
                    pd.to_numeric(series, errors="coerce"), unit="D")
                if parsed.notna().sum() >= 0.85 * len(series.dropna()):
                    return parsed, "datetime (excel serial)"
        except Exception:
            pass
        return None, None

    import warnings
    for kwargs, label in [
        ({"errors": "coerce"}, "datetime"),
        ({"errors": "coerce", "dayfirst": True}, "datetime (dayfirst)"),
        ({"errors": "coerce", "format": "mixed"}, "datetime (mixed)"),
    ]:
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                parsed_sample = pd.to_datetime(sample, **kwargs)
                if (parsed_sample.notna().sum() >= 0.85 * len(sample)
                        and parsed_sample.dropna().nunique() >= 2):
                    full = pd.to_datetime(series, **kwargs)
                    return full, label
        except (ValueError, TypeError):
            continue
    return None, None


def _detect_pii_columns(df: pd.DataFrame) -> dict:
    """Scan object columns for PII patterns. Returns {col: pii_type}.

    Triggers if >=20% of non-null values match the pattern — high enough to
    avoid false positives, low enough to catch real PII columns that have
    some blank entries.
    """
    out: dict = {}
    for col in df.select_dtypes(include=["object"]).columns:
        nn = df[col].dropna().astype(str).head(500)
        if nn.empty:
            continue
        for kind, pat in _PII_PATTERNS.items():
            if (nn.str.contains(pat).mean() >= 0.2):
                out[col] = kind
                break
    return out


def _detect_placeholder_numbers(series: pd.Series) -> int | None:
    """Find a numeric placeholder value if one dominates the column.

    Returns the placeholder value (as int) if a known sentinel value
    accounts for >=2% of the column AND occurs as a clear outlier compared
    to the rest of the distribution.
    """
    if not pd.api.types.is_numeric_dtype(series):
        return None
    nn = series.dropna()
    if len(nn) < 50:
        return None
    counts = nn.value_counts()
    for ph in _PLACEHOLDER_NUMS:
        if ph in counts and counts[ph] / len(nn) >= 0.02:
            # Heuristic: the placeholder should sit far above (or below) the
            # rest of the distribution — not just a frequent legitimate value.
            rest = nn[nn != ph]
            if rest.empty:
                return int(ph)
            q1, q3 = rest.quantile(0.25), rest.quantile(0.75)
            iqr = q3 - q1
            if ph > q3 + 3 * iqr or ph < q1 - 3 * iqr:
                return int(ph)
    return None


def clean_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Clean the dataframe *conservatively*. Guiding principles:

    - Never fabricate values. NaN stays NaN. We trust downstream analyses
      (pandas aggregations skip NaN) to handle missingness correctly.
    - Never drop columns for being mostly-null — high-null columns are often
      the most analytically interesting (discounts, returns, etc.). Report
      them instead so the AI narrative can reason about them.
    - Deduplicate at the right grain. If there's an ID-like primary key,
      dedupe on that. Otherwise dedupe on full rows.
    - Canonicalize categorical spelling *before* any chart is built.
    """
    summary = {
        "original_shape": list(df.shape),
        "whitespace_columns_fixed": [],
        "duplicates_removed": 0,
        "duplicate_key": None,
        "id_columns": [],
        "types_inferred": {},
        "category_merges": {},           # {col: {variant: canonical}}
        "high_null_columns": {},         # {col: pct}
        "negative_in_positive_cols": {}, # {col: count}
        "zero_in_positive_cols": {},     # {col: count}
        "placeholder_values_nulled": {}, # {col: {value: count}}
        "pii_columns": {},               # {col: kind}
        "rows_with_any_null": 0,
    }

    # 1. Trim column names + dedupe header collisions.
    new_cols = {c: str(c).strip() for c in df.columns}
    changed = [c for c, nc in new_cols.items() if c != nc]
    df = df.rename(columns=new_cols)
    summary["whitespace_columns_fixed"] = changed

    # 2. Strip string whitespace + replace expanded list of empty sentinels.
    # Senior-analyst heuristic: more sentinels than just "N/A". CSV exports
    # from Snowflake, Looker, GA, etc. each pick a different "missing" token.
    SENTINELS = {"nan", "None", "NaN", "NULL", "null", "Null", "N/A", "n/a",
                 "NA", "na", "-", "--", "?", "unknown", "Unknown", "UNKNOWN",
                 "(blank)", "(empty)", "<NA>", "#N/A", ""}
    for col in df.select_dtypes(include=["object"]).columns:
        s = df[col].astype(str).str.strip()
        df[col] = s.replace({v: np.nan for v in SENTINELS})

    # 3. Type inference — booleans, dates, numerics. Smart parsers handle
    # currency, EU decimals, accounting parens, Excel serials, dayfirst, etc.
    for col in df.columns:
        if df[col].dtype != object:
            continue
        sample = df[col].dropna().astype(str).head(300)
        if len(sample) == 0:
            continue

        # Boolean (expanded: y/n, t/f, 0/1, true/false).
        lower_vals = set(sample.str.lower().str.strip().unique())
        bool_textual = {"true", "false", "yes", "no", "y", "n", "t", "f",
                        "1", "0"}
        if lower_vals.issubset(bool_textual) and 1 <= len(lower_vals) <= 2:
            df[col] = df[col].astype(str).str.lower().str.strip().map(
                {"true": True, "false": False, "yes": True, "no": False,
                 "y": True, "n": False, "t": True, "f": False,
                 "1": True, "0": False}
            )
            summary["types_inferred"][col] = "boolean"
            continue

        # Date — try a few strategies (ISO, dayfirst, mixed, Excel serial).
        parsed_dt, dt_label = _parse_dates_smart(df[col])
        if parsed_dt is not None:
            df[col] = parsed_dt
            summary["types_inferred"][col] = dt_label
            continue

        # Numeric — currency-aware, parens-as-negative, EU decimal, percent.
        parsed_num, num_label = _parse_numeric_smart(df[col])
        if num_label:
            df[col] = parsed_num
            summary["types_inferred"][col] = num_label

    # 4. Canonicalize categorical spelling — fuzzy merge near-duplicate
    # labels ("bengaluru"/"Bangalore"/"bangalore", "hyd"/"hyderabad"/"hyderbad")
    # for low-cardinality object columns. Applied BEFORE any chart so every
    # visualization sees merged labels.
    for col in df.select_dtypes(include=["object"]).columns:
        s = df[col]
        nn = s.dropna()
        if len(nn) == 0:
            continue
        nunique = nn.nunique()
        if 2 <= nunique <= max(40, int(len(nn) * 0.1)):
            merged, audit = _fuzzy_canonicalize(nn, cutoff=0.85)
            if audit:
                df.loc[nn.index, col] = merged
                summary["category_merges"][col] = audit

    # 5. Dedupe — ID-aware. If a primary-key-like column exists and has
    # repeats, dedupe on it; otherwise dedupe full rows.
    id_cols = _detect_id_columns(df)
    summary["id_columns"] = id_cols
    before = len(df)
    if id_cols:
        key = id_cols[0]
        if df[key].duplicated().any():
            df = df.drop_duplicates(subset=[key], keep="first").reset_index(drop=True)
            summary["duplicate_key"] = key
    if len(df) == before:
        df = df.drop_duplicates().reset_index(drop=True)
        if not id_cols:
            summary["duplicate_key"] = "(full-row)"
    summary["duplicates_removed"] = before - len(df)

    # 6. Nullify clearly-invalid values. A zero in a column called
    #    "Revenue" is almost always missing data, not a $0 sale — treating
    #    it as real corrupts every aggregate (drags mean to 0, flips
    #    rankings, inflates "no significant difference" false negatives).
    #    Same reasoning for negatives in columns that must be positive,
    #    and for rate columns outside [0, 1] or [0, 100].
    AMOUNT_POSITIVE_HINTS = (
        "price", "revenue", "sales", "amount", "qty", "quantity",
        "units", "mrp", "cost", "spend", "total", "gmv", "value",
    )
    RATE_HINTS = ("discount", "rate", "ratio", "pct", "percent", "conversion", "ctr")

    summary["suspect_zeros_nulled"] = {}
    summary["suspect_negatives_nulled"] = {}
    summary["invalid_rates_nulled"] = {}

    n = len(df)
    for col in df.columns:
        if not pd.api.types.is_numeric_dtype(df[col]):
            continue
        name_l = str(col).lower()
        is_amount = any(h in name_l for h in AMOUNT_POSITIVE_HINTS)
        is_rate = any(h in name_l for h in RATE_HINTS)

        if is_amount:
            neg_mask = df[col] < 0
            zer_mask = df[col] == 0
            n_neg = int(neg_mask.sum())
            # Only treat zeros as missing if they're a meaningful
            # fraction — a legitimate $0 row here and there should stay.
            z_frac = (zer_mask.sum() / max(1, n))
            n_zer = int(zer_mask.sum()) if z_frac >= 0.01 else 0
            if n_neg:
                summary["negative_in_positive_cols"][col] = n_neg
                summary["suspect_negatives_nulled"][col] = n_neg
                df.loc[neg_mask, col] = np.nan
            if n_zer:
                summary["zero_in_positive_cols"][col] = n_zer
                summary["suspect_zeros_nulled"][col] = n_zer
                df.loc[zer_mask, col] = np.nan

        if is_rate:
            col_nn = df[col].dropna()
            if len(col_nn):
                # Decide scale: if max <= 1.5 we treat as 0–1 fraction,
                # otherwise as 0–100 percent.
                scale_max = 1.0 if col_nn.max() <= 1.5 else 100.0
                invalid_mask = (df[col] < 0) | (df[col] > scale_max)
                n_inv = int(invalid_mask.sum())
                if n_inv:
                    summary["invalid_rates_nulled"][col] = n_inv
                    df.loc[invalid_mask, col] = np.nan

    # 7. Reconstruct derived amount columns. If a column looks like
    #    revenue/total but has lots of null values, and we can find
    #    plausible price/quantity (and optional discount) components,
    #    recompute the missing rows: revenue = price * qty * (1 - discount)
    summary["revenue_reconstructed"] = {}

    def _match_col(keywords: tuple[str, ...]) -> str | None:
        for c in df.columns:
            cl = str(c).lower()
            if any(kw in cl for kw in keywords) and pd.api.types.is_numeric_dtype(df[c]):
                return c
        return None

    revenue_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])
                    and any(h in str(c).lower() for h in ("revenue", "sales", "total", "gmv"))]
    price_col = _match_col(("price", "mrp", "unit_price", "list_price"))
    qty_col = _match_col(("units", "qty", "quantity"))
    discount_col = _match_col(("discount",))

    if revenue_cols and price_col and qty_col:
        for rev in revenue_cols:
            missing_mask = df[rev].isna()
            if not missing_mask.any():
                continue
            comp = df[[price_col, qty_col]].copy()
            usable = comp[price_col].notna() & comp[qty_col].notna() & missing_mask
            if not usable.any():
                continue
            d_factor = 1.0
            if discount_col is not None:
                d = df[discount_col].copy()
                d_scale = 1.0 if d.dropna().max() <= 1.5 else 100.0
                d_factor = (1.0 - (d.fillna(0) / d_scale))
            recomputed = df[price_col] * df[qty_col]
            if isinstance(d_factor, pd.Series):
                recomputed = recomputed * d_factor
            df.loc[usable, rev] = recomputed[usable]
            summary["revenue_reconstructed"][rev] = {
                "rows_filled": int(usable.sum()),
                "formula": (
                    f"{price_col} * {qty_col}"
                    + (f" * (1 - {discount_col}/{int(100 if discount_col and df[discount_col].dropna().max() > 1.5 else 1)})"
                       if discount_col is not None else "")
                ),
            }

    # 8. Detect and nullify analyst-placeholder numbers (9999, -1, etc.)
    #    that are obviously sentinels rather than real measurements.
    for col in df.columns:
        ph = _detect_placeholder_numbers(df[col])
        if ph is not None:
            mask = df[col] == ph
            count = int(mask.sum())
            if count:
                df.loc[mask, col] = np.nan
                summary["placeholder_values_nulled"][col] = {str(ph): count}

    # 9. PII detection — flag columns that contain emails, phone numbers,
    #    card numbers, or SSNs. We don't redact (analysts need the column
    #    to dedupe customers, etc.) — we just surface the finding.
    summary["pii_columns"] = _detect_pii_columns(df)

    # 10. Record — don't fix — high-null columns.
    for col in df.columns:
        null_pct = (df[col].isna().sum() / n) if n else 0.0
        if null_pct >= 0.30:
            summary["high_null_columns"][col] = round(float(null_pct), 3)

    summary["rows_with_any_null"] = int(df.isna().any(axis=1).sum())
    summary["cleaned_shape"] = list(df.shape)
    return df, summary


# -----------------------------------------------------------------------------
# Dataset profiling
# -----------------------------------------------------------------------------
def profile_dataframe(df: pd.DataFrame) -> dict:
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    date_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64"]).columns.tolist()
    bool_cols = df.select_dtypes(include=["bool"]).columns.tolist()
    cat_cols = [c for c in df.columns if c not in numeric_cols + date_cols + bool_cols]

    return {
        "shape": list(df.shape),
        "columns": list(df.columns),
        "dtypes": {c: str(df[c].dtype) for c in df.columns},
        "numeric_cols": numeric_cols,
        "date_cols": date_cols,
        "bool_cols": bool_cols,
        "categorical_cols": cat_cols,
        "null_counts": {c: int(df[c].isna().sum()) for c in df.columns},
        "nunique": {c: int(df[c].nunique()) for c in df.columns},
        "sample_rows": df.head(5).astype(str).to_dict(orient="records"),
        "describe_numeric": (
            df[numeric_cols].describe().round(4).to_dict() if numeric_cols else {}
        ),
    }


# -----------------------------------------------------------------------------
# Domain templates
# -----------------------------------------------------------------------------
TEMPLATES = {
    "email": {
        "name": "Email Marketing",
        "kpis": "open rate, click-through rate, unsubscribe rate, bounce rate, send volume, engagement over time, subject line performance",
        "guidance": "Focus on deliverability and engagement. Look for time-of-day or subject-line patterns. Compute rates as percentages when possible.",
    },
    "campaign": {
        "name": "Campaign Performance",
        "kpis": "impressions, clicks, CTR, conversions, cost per acquisition, ROAS, channel comparison, trend over time",
        "guidance": "Benchmark channels against each other. Surface efficiency metrics (CPA, ROAS) and trend shifts.",
    },
    "abtest": {
        "name": "A/B Testing",
        "kpis": "variant comparison, statistical significance, conversion lift, sample size adequacy, confidence intervals, winner recommendation",
        "guidance": "Run significance tests (chi-square or t-test) and state winner with confidence level. Call out if sample size is too small.",
    },
    "site": {
        "name": "Site Usage",
        "kpis": "sessions, bounce rate, pages per session, top pages, traffic sources, device breakdown, funnel drop-off",
        "guidance": "Identify top and bottom pages, device/source segments, and any funnel leaks.",
    },
    "sales": {
        "name": "Sales Performance",
        "kpis": "revenue by region or product, order volume, average order value, growth rate, top performers, period-over-period comparison",
        "guidance": "Compare segments and time periods. Highlight outperformers and laggards.",
    },
    "benchmark": {
        "name": "Benchmark Survey",
        "kpis": "response distribution, average scores by category, top/bottom performing segments, trend comparison",
        "guidance": "Summarize distributions and compare segments. If time data exists, show trend.",
    },
    "general": {
        "name": "General All-Inclusive Analysis",
        "kpis": "all relevant trends, distributions, segment comparisons, and notable relationships",
        "guidance": "Explore the dataset broadly. Pick the 6-10 most informative charts.",
    },
}


# -----------------------------------------------------------------------------
# AI provider abstraction
# -----------------------------------------------------------------------------
ANALYSIS_PROMPT = """You are a SENIOR B2C DATA ANALYST with 10+ years of experience.
You are NOT a generic AI assistant. You write like an analyst presenting to
the CEO: direct, specific, recommendation-driven, free of filler. Return ONLY
valid JSON.

DETECTED DATASET ARCHETYPE: {archetype_name} (confidence {archetype_conf})
ARCHETYPE DESCRIPTION: {archetype_desc}
ROLE COLUMNS DETECTED: {role_columns}

The right vocabulary for this archetype:
{vocab}

PRE-COMPUTED PLAYBOOK OUTPUT (this is the ground truth — every number you
cite must come from here or from the stats block below):
{playbook_json}

PRE-COMPUTED GROUND-TRUTH STATS:
{stats_json}

DATASET PROFILE (schema reference only — do not infer business meaning that
contradicts the archetype):
{profile_json}

CLEANING REPORT (honest record of what was and wasn't done to the data):
{clean_json}

ANALYSIS MODE: {mode}
DOMAIN GUIDANCE: {guidance}
CUSTOM USER QUESTIONS: {custom}
BENCHMARKS: {benchmarks}

Return a JSON object with EXACTLY this schema:
{{
  "executive_summary": "4-6 sentences in senior-analyst voice. Lead with the most important finding (not a description). Reference the archetype's vocabulary. Cite specific numbers from the playbook output. End with one concrete recommendation.",
  "kpi_cards": [
    {{"label": "string", "value": "string (formatted)", "subtext": "string"}}
  ],
  "data_quality_notes": ["observation 1", "observation 2"],
  "followup_questions": ["question 1", "question 2", "question 3"],
  "sql_queries": [
    {{"title": "Query description", "sql": "SELECT ... (Snowflake compatible)"}}
  ]
}}

CRITICAL ANALYST RULES — break any of these and the response is wrong:
1. NO generic observations. "Revenue trends upward" or "the data shows
   variation" are banned. Be specific: which segment, which channel, by
   what amount, vs what baseline.
2. Use the archetype's vocabulary. For orders: AOV, repeat rate, RFM tiers,
   month-N retention, top-decile share. For marketing: ROAS, CAC, CTR/CVR
   by channel, blended vs paid. Don't say "high-value customers" if you
   mean "Champions" (the RFM segment).
3. Every number you cite must appear in the playbook output or stats above.
   No invented totals, no rounded-feel guesses.
4. Prefer the playbook's KPI list and segments for kpi_cards — they're
   already correct. Reformat them, don't replace them.
5. The recommendation in executive_summary must be a SPECIFIC action a
   manager could take this quarter — not "consider improving X".
6. In data_quality_notes, surface the playbook 'alerts' and the cleaning
   findings (high_null_columns, category_merges, suspect_negatives_nulled,
   PII columns, etc.). Honesty about caveats matters.
7. Charts are built deterministically in code — do NOT propose charts.
8. SQL must be Snowflake-compatible and assume the table is named `dataset`.
9. Respond with ONLY the JSON object, no markdown fences.
"""


ARCHETYPE_VOCAB = {
    "orders": (
        "AOV (average order value), repeat-purchase rate, RFM segmentation "
        "(Champions / Loyal / At Risk / Hibernating / New / Need Attention), "
        "cohort retention by month-offset, top-decile revenue concentration, "
        "Pareto share, gross margin, return rate, discount lift."
    ),
    "customers": (
        "RFM scoring, segment share of revenue, CLV (customer lifetime value), "
        "predicted churn risk, acquisition-channel mix, tenure cohorts."
    ),
    "marketing": (
        "ROAS (revenue ÷ spend), CAC (cost per acquisition), CTR (clicks ÷ "
        "impressions), CVR (conversions ÷ clicks), blended vs paid, channel "
        "mix, frequency-response curve, diminishing returns, attribution window."
    ),
    "sessions": (
        "funnel drop-off by step, bounce rate, session duration distribution, "
        "source/medium mix, top exit pages, conversion rate by traffic source."
    ),
    "subscriptions": (
        "MRR (monthly recurring revenue), gross/net churn, cohort survival "
        "curves, save rate, expansion vs contraction, LTV by cohort, "
        "downgrade/upgrade flows."
    ),
    "reviews": (
        "sentiment distribution, rating by SKU, low-rating topic clusters, "
        "review velocity over time, verified vs unverified, NPS bucketing "
        "(promoter/passive/detractor)."
    ),
    "catalog": (
        "margin by category, stockout risk, long-tail revenue share, "
        "seasonal SKUs, price-elasticity proxies, assortment depth."
    ),
    "generic": (
        "Standard exploratory data analysis — distributions, top categories, "
        "correlations, time-series trends. Use plain analyst language."
    ),
}


def _safe_json_extract(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("No JSON object found in model response")
    return json.loads(text[start:end + 1])


def _build_prompt(profile: dict, mode: str, custom: str, benchmarks: list,
                  grounded: dict | None = None,
                  clean_summary: dict | None = None,
                  archetype: object | None = None,
                  playbook: dict | None = None) -> str:
    tpl = TEMPLATES.get(mode, TEMPLATES["general"])
    arch_name = getattr(archetype, "name", "generic")
    arch_conf = getattr(archetype, "confidence", 0.0)
    arch_desc = arch_mod.archetype_description(arch_name)
    role_cols = getattr(archetype, "role_columns", {}) or {}
    vocab = ARCHETYPE_VOCAB.get(arch_name, ARCHETYPE_VOCAB["generic"])

    return ANALYSIS_PROMPT.format(
        archetype_name=arch_name,
        archetype_conf=arch_conf,
        archetype_desc=arch_desc,
        role_columns=json.dumps(role_cols, default=str)[:1000],
        vocab=vocab,
        playbook_json=json.dumps(playbook or {}, default=str)[:6000],
        profile_json=json.dumps(profile, default=str)[:6000],
        stats_json=json.dumps(grounded or {}, default=str)[:6000],
        clean_json=json.dumps(clean_summary or {}, default=str)[:4000],
        mode=tpl["name"],
        guidance=tpl["guidance"],
        custom=custom or "None",
        benchmarks=json.dumps(benchmarks) if benchmarks else "None",
    )


def call_openai(api_key: str, prompt: str, strict: bool = False) -> str:
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    system = "You are a data analytics expert. Return only valid JSON."
    if strict:
        system += " Your previous response was not valid JSON. Return ONLY a JSON object with no other text, no markdown, no fences."
    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
        response_format={"type": "json_object"},
    )
    return resp.choices[0].message.content


def call_anthropic(api_key: str, prompt: str, strict: bool = False) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    system = "You are a data analytics expert. Return only valid JSON, no markdown fences, no prose."
    if strict:
        system += " CRITICAL: Return ONLY a JSON object starting with { and ending with }. No other text."
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=system,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.content[0].text


def call_gemini(api_key: str, prompt: str, strict: bool = False) -> str:
    import google.generativeai as genai
    genai.configure(api_key=api_key)
    system = "You are a data analytics expert. Return only valid JSON with no markdown fences."
    if strict:
        system += " Return ONLY a JSON object. No prose. No code fences."
    model = genai.GenerativeModel(
        "gemini-1.5-pro",
        system_instruction=system,
        generation_config={"response_mime_type": "application/json", "temperature": 0.2},
    )
    resp = model.generate_content(prompt)
    return resp.text


def analyze(provider: str, api_key: str, profile: dict, mode: str,
            custom: str, benchmarks: list,
            grounded: dict | None = None,
            clean_summary: dict | None = None,
            archetype: object | None = None,
            playbook: dict | None = None) -> dict:
    """Unified AI provider call with one retry on malformed JSON."""
    prompt = _build_prompt(profile, mode, custom, benchmarks, grounded,
                            clean_summary, archetype=archetype, playbook=playbook)
    caller = {"openai": call_openai, "anthropic": call_anthropic, "gemini": call_gemini}.get(provider)
    if not caller:
        raise ValueError(f"Unknown provider: {provider}")

    try:
        raw = caller(api_key, prompt, strict=False)
        return _safe_json_extract(raw)
    except (json.JSONDecodeError, ValueError) as e:
        log.warning(f"First AI call produced invalid JSON ({e}); retrying with stricter prompt.")
        raw = caller(api_key, prompt, strict=True)
        return _safe_json_extract(raw)


# -----------------------------------------------------------------------------
# Chart builders
# -----------------------------------------------------------------------------
def _fig_layout(fig, title):
    fig.update_layout(
        title=dict(
            text=title,
            font=dict(size=15, family="Inter, -apple-system, system-ui, sans-serif", color="#141815"),
            x=0, xanchor="left", pad=dict(l=6),
        ),
        font=dict(family="Inter, -apple-system, system-ui, sans-serif", color="#141815", size=12),
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        margin=dict(l=50, r=30, t=50, b=50),
        colorway=PALETTE,
        legend=dict(bgcolor="rgba(255,255,255,0)", bordercolor="#E2E7E2", borderwidth=0, font=dict(size=11)),
        hoverlabel=dict(bgcolor="#141815", font=dict(color="#F6F7F4", family="Inter")),
    )
    fig.update_xaxes(showgrid=True, gridcolor="#EDEFE9", zeroline=False, linecolor="#E2E7E2")
    fig.update_yaxes(showgrid=True, gridcolor="#EDEFE9", zeroline=False, linecolor="#E2E7E2")
    return fig


def build_chart(df: pd.DataFrame, spec: dict) -> dict | None:
    """Convert an AI chart spec into a Plotly figure dict. Returns None if columns missing."""
    try:
        ctype = (spec.get("chart_type") or "bar").lower()
        title = spec.get("title") or "Chart"
        x = spec.get("x")
        y = spec.get("y")
        color = spec.get("color")
        agg = (spec.get("agg") or "none").lower()

        cols = [c for c in [x, y, color] if c]
        missing = [c for c in cols if c and c not in df.columns]
        if missing:
            log.warning(f"Skipping chart '{title}': missing columns {missing}")
            return None

        d = df.copy()
        if agg in ("sum", "mean", "count", "median") and x and y and y in df.columns:
            group_cols = [x] + ([color] if color and color != x else [])
            if agg == "count":
                d = d.groupby(group_cols, dropna=False).size().reset_index(name=y)
            else:
                d = d.groupby(group_cols, dropna=False)[y].agg(agg).reset_index()

        if ctype == "bar":
            fig = px.bar(d, x=x, y=y, color=color)
        elif ctype == "line":
            if x and pd.api.types.is_datetime64_any_dtype(d[x]):
                d = d.sort_values(x)
            fig = px.line(d, x=x, y=y, color=color, markers=True)
        elif ctype == "pie":
            fig = px.pie(d, names=x, values=y if y else None)
        elif ctype == "scatter":
            fig = px.scatter(d, x=x, y=y, color=color, opacity=0.75)
        elif ctype == "histogram":
            fig = px.histogram(d, x=x, color=color)
        elif ctype == "box":
            fig = px.box(d, x=x, y=y, color=color)
        elif ctype == "area":
            fig = px.area(d, x=x, y=y, color=color)
        elif ctype == "heatmap":
            if x and y:
                pivot = df.pivot_table(index=y, columns=x, aggfunc="size", fill_value=0)
                fig = px.imshow(pivot, color_continuous_scale=[[0, "#ECFDF5"], [0.5, "#22C55E"], [1, "#14532D"]])
            else:
                return None
        else:
            fig = px.bar(d, x=x, y=y, color=color)

        fig = _fig_layout(fig, title)
        return json.loads(json.dumps(fig.to_dict(), cls=PlotlyJSONEncoder))
    except Exception as e:
        log.warning(f"Chart build failed for '{spec.get('title')}': {e}")
        return None


AMOUNT_HINTS = ("revenue", "sales", "amount", "profit", "spend", "cost",
                "price", "total", "value", "gmv")
COUNT_HINTS = ("qty", "quantity", "units", "count", "orders", "sessions",
               "users", "clicks", "impressions", "views")


def _is_amount_col(name: str) -> bool:
    n = str(name).lower()
    return any(h in n for h in AMOUNT_HINTS)


def _is_count_col(name: str) -> bool:
    n = str(name).lower()
    return any(h in n for h in COUNT_HINTS)


def _to_fig_dict(fig) -> dict:
    return json.loads(json.dumps(fig.to_dict(), cls=PlotlyJSONEncoder))


def build_auto_charts(df: pd.DataFrame) -> list[dict]:
    """Build a curated, code-deterministic set of charts.

    Guarantees:
    - Time series plots COUNT of rows per period (never raw id values).
    - Categorical breakdowns use horizontal bar of mean/median or box plots,
      not stacked histograms with 20 colors.
    - IDs and obvious key columns are excluded from numeric summarization.
    - NaN rows are excluded per-chart (not globally), preserving each
      column's legitimate sample size.
    """
    charts: list[dict] = []
    date_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64"]).columns.tolist()
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    bool_cols = df.select_dtypes(include=["bool"]).columns.tolist()
    id_cols = set(_detect_id_columns(df))
    # Exclude ID columns from numeric summarization (they're not metrics).
    measure_cols = [c for c in numeric_cols if c not in id_cols]
    # Amount-like and count-like measures get preferential treatment.
    amount_cols = [c for c in measure_cols if _is_amount_col(c)]
    count_cols = [c for c in measure_cols if _is_count_col(c)]

    # Low-cardinality categoricals = good breakdown dimensions.
    cat_cols = []
    for c in df.columns:
        if c in id_cols:
            continue
        if df[c].dtype == object or c in bool_cols:
            nn = df[c].dropna()
            if 2 <= nn.nunique() <= 30:
                cat_cols.append(c)

    # ---- 1. Orders per month (count of rows) -------------------------------
    if date_cols:
        dcol = date_cols[0]
        d = df[[dcol]].dropna().copy()
        if len(d) >= 5:
            d["_period"] = d[dcol].dt.to_period("M").dt.to_timestamp()
            g = d.groupby("_period").size().reset_index(name="Rows")
            fig = px.bar(g, x="_period", y="Rows")
            fig.update_traces(marker_color=PALETTE[0])
            fig = _fig_layout(fig, f"Row volume per month — based on {dcol}")
            fig.update_xaxes(title_text=dcol)
            fig.update_yaxes(title_text="Count of rows")
            charts.append({
                "title": f"Row volume per month ({dcol})",
                "insight": (
                    f"Counts the number of rows falling in each month of "
                    f"{dcol}. Range: {d[dcol].min().date()} to "
                    f"{d[dcol].max().date()}. Peak month: "
                    f"{g.loc[g['Rows'].idxmax(), '_period'].strftime('%Y-%m')} "
                    f"with {int(g['Rows'].max())} rows."
                ),
                "figure": _to_fig_dict(fig),
            })

    # ---- 2. Primary amount metric over time (sum per month) ----------------
    if date_cols and amount_cols:
        dcol = date_cols[0]
        acol = amount_cols[0]
        d = df[[dcol, acol]].dropna().copy()
        if len(d) >= 5:
            d["_period"] = d[dcol].dt.to_period("M").dt.to_timestamp()
            g = d.groupby("_period")[acol].sum().reset_index()
            fig = px.line(g, x="_period", y=acol, markers=True)
            fig.update_traces(line=dict(color=PALETTE[0], width=2.5))
            fig = _fig_layout(fig, f"{acol} over time (monthly sum)")
            fig.update_xaxes(title_text=dcol)
            fig.update_yaxes(title_text=f"Sum of {acol}")
            top_period = g.loc[g[acol].idxmax(), "_period"].strftime("%Y-%m")
            charts.append({
                "title": f"{acol} over time",
                "insight": (
                    f"Monthly sum of {acol}. Total across range: "
                    f"{g[acol].sum():,.2f}. Peak month: {top_period} "
                    f"({g[acol].max():,.2f})."
                ),
                "figure": _to_fig_dict(fig),
            })

    # ---- 3. Top-N bar: primary amount by each low-card categorical ---------
    primary_amount = amount_cols[0] if amount_cols else (count_cols[0] if count_cols else None)
    for cat in cat_cols[:4]:
        if primary_amount is None:
            # Fall back to simple counts.
            g = df[cat].value_counts(dropna=True).head(15).reset_index()
            g.columns = [cat, "Rows"]
            fig = px.bar(g, x="Rows", y=cat, orientation="h", color=cat,
                         color_discrete_sequence=CAT_PALETTE)
            fig = _fig_layout(fig, f"Row count by {cat}")
            fig.update_layout(yaxis={"categoryorder": "total ascending"}, showlegend=False)
            charts.append({
                "title": f"Row count by {cat}",
                "insight": f"Counts per {cat}. Top: {g.iloc[0][cat]} ({int(g.iloc[0]['Rows'])} rows).",
                "figure": _to_fig_dict(fig),
            })
            continue
        d = df[[cat, primary_amount]].dropna()
        if len(d) < 3 or d[cat].nunique() < 2:
            continue
        g = d.groupby(cat)[primary_amount].agg(["sum", "mean", "count"]).reset_index()
        g = g.sort_values("sum", ascending=False).head(15)
        fig = px.bar(g, x="sum", y=cat, orientation="h", color=cat,
                     color_discrete_sequence=CAT_PALETTE,
                     hover_data={"mean": ":,.2f", "count": True})
        fig = _fig_layout(fig, f"Total {primary_amount} by {cat}")
        fig.update_layout(yaxis={"categoryorder": "total ascending"}, showlegend=False)
        fig.update_xaxes(title_text=f"Sum of {primary_amount}")
        top = g.iloc[0]
        charts.append({
            "title": f"{primary_amount} by {cat}",
            "insight": (
                f"Top {cat}: {top[cat]} — total {top['sum']:,.2f} "
                f"(mean {top['mean']:,.2f} across {int(top['count'])} rows)."
            ),
            "figure": _to_fig_dict(fig),
        })

    # ---- 4. Box plot: primary amount distribution by top categorical -------
    if primary_amount and cat_cols:
        cat = cat_cols[0]
        d = df[[cat, primary_amount]].dropna()
        if d[cat].nunique() <= 20 and len(d) >= 10:
            fig = px.box(d, x=cat, y=primary_amount, color=cat,
                         color_discrete_sequence=CAT_PALETTE, points="outliers")
            fig = _fig_layout(fig, f"{primary_amount} distribution by {cat}")
            fig.update_layout(showlegend=False)
            charts.append({
                "title": f"{primary_amount} distribution by {cat}",
                "insight": (
                    f"Box plot of {primary_amount} for each {cat}. Compare "
                    f"medians and spread — wider boxes indicate higher "
                    f"variance within that {cat}."
                ),
                "figure": _to_fig_dict(fig),
            })

    # ---- 4b. Binary: discount vs no-discount comparison --------------------
    # The clearest analytical question around a discount column is not "does
    # revenue correlate with discount %" (noisy, confounded) but "do rows
    # that received ANY discount differ from rows that received none?".
    discount_col = next(
        (c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])
         and "discount" in str(c).lower()),
        None,
    )
    if discount_col and primary_amount and primary_amount != discount_col:
        d = df[[discount_col, primary_amount]].dropna()
        if len(d) >= 20:
            has_disc = (d[discount_col] > 0).map({True: "With discount", False: "No discount"})
            g = d.assign(_grp=has_disc).groupby("_grp")[primary_amount].agg(
                ["mean", "median", "count", "sum"]
            ).reset_index()
            if len(g) == 2:
                fig = px.bar(
                    g, x="_grp", y="mean", color="_grp",
                    color_discrete_sequence=[CAT_PALETTE[0], CAT_PALETTE[1]],
                    hover_data={"median": ":,.2f", "count": True, "sum": ":,.2f"},
                )
                fig = _fig_layout(fig, f"Mean {primary_amount}: discount vs no-discount")
                fig.update_layout(showlegend=False)
                fig.update_xaxes(title_text="")
                fig.update_yaxes(title_text=f"Mean {primary_amount}")
                with_row = g[g["_grp"] == "With discount"].iloc[0]
                no_row = g[g["_grp"] == "No discount"].iloc[0]
                lift = (with_row["mean"] - no_row["mean"]) / no_row["mean"] * 100 if no_row["mean"] else 0
                charts.append({
                    "title": f"Discount impact on {primary_amount}",
                    "insight": (
                        f"With discount (n={int(with_row['count'])}): mean "
                        f"{with_row['mean']:,.2f}. No discount "
                        f"(n={int(no_row['count'])}): mean {no_row['mean']:,.2f}. "
                        f"Difference: {lift:+.1f}%."
                    ),
                    "figure": _to_fig_dict(fig),
                })

    # ---- 4c. Profit-by-group (in addition to revenue/primary amount) -------
    # If a profit/margin column exists, surface it separately — aggregating
    # revenue hides whether the revenue was profitable.
    profit_col = next(
        (c for c in measure_cols
         if any(h in str(c).lower() for h in ("profit", "margin", "gross"))),
        None,
    )
    if profit_col and profit_col != primary_amount and cat_cols:
        for cat in cat_cols[:2]:
            d = df[[cat, profit_col]].dropna()
            if len(d) < 3 or d[cat].nunique() < 2:
                continue
            g = d.groupby(cat)[profit_col].agg(["sum", "mean", "count"]).reset_index()
            g = g.sort_values("sum", ascending=False).head(15)
            fig = px.bar(g, x="sum", y=cat, orientation="h", color=cat,
                         color_discrete_sequence=CAT_PALETTE,
                         hover_data={"mean": ":,.2f", "count": True})
            fig = _fig_layout(fig, f"Total {profit_col} by {cat}")
            fig.update_layout(yaxis={"categoryorder": "total ascending"}, showlegend=False)
            fig.update_xaxes(title_text=f"Sum of {profit_col}")
            top = g.iloc[0]
            charts.append({
                "title": f"{profit_col} by {cat}",
                "insight": (
                    f"Top {cat} by {profit_col}: {top[cat]} — total "
                    f"{top['sum']:,.2f} (mean {top['mean']:,.2f} across "
                    f"{int(top['count'])} rows)."
                ),
                "figure": _to_fig_dict(fig),
            })

    # ---- 5. Scatter: the two numerics with highest abs correlation ---------
    if len(measure_cols) >= 2:
        d = df[measure_cols].dropna()
        if len(d) >= 10:
            corr = d.corr(numeric_only=True).abs()
            np.fill_diagonal(corr.values, 0)
            if corr.values.size and np.isfinite(corr.values).any():
                idx = np.unravel_index(np.nanargmax(corr.values), corr.shape)
                xcol, ycol = corr.columns[idx[0]], corr.columns[idx[1]]
                r_val = df[[xcol, ycol]].corr().iloc[0, 1]
                fig = px.scatter(df.dropna(subset=[xcol, ycol]),
                                 x=xcol, y=ycol, opacity=0.6,
                                 color_discrete_sequence=[PALETTE[0]],
                                 trendline=None)
                fig = _fig_layout(fig, f"{ycol} vs {xcol} (r = {r_val:.3f})")
                charts.append({
                    "title": f"{ycol} vs {xcol}",
                    "insight": (
                        f"Pearson r = {r_val:.3f} between {xcol} and {ycol}. "
                        f"Values near 0 indicate no linear relationship; "
                        f"values near ±1 indicate strong linear relationship."
                    ),
                    "figure": _to_fig_dict(fig),
                })

    return charts


def correlation_heatmap(df: pd.DataFrame) -> dict | None:
    numeric = df.select_dtypes(include=[np.number])
    # Exclude ID-like columns — correlations with an auto-incrementing
    # primary key are meaningless and crowd the heatmap.
    id_cols = set(_detect_id_columns(df))
    numeric = numeric[[c for c in numeric.columns if c not in id_cols]]
    if numeric.shape[1] < 2:
        return None
    corr = numeric.corr().round(3)
    fig = px.imshow(
        corr, text_auto=True, color_continuous_scale="RdBu_r",
        zmin=-1, zmax=1, aspect="auto"
    )
    fig = _fig_layout(fig, "Correlation Heatmap (Numeric Columns)")
    return json.loads(json.dumps(fig.to_dict(), cls=PlotlyJSONEncoder))


def time_series_trend(df: pd.DataFrame) -> dict | None:
    """Row volume per week using the first date column. Guaranteed count,
    never id values."""
    date_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64"]).columns.tolist()
    if not date_cols:
        return None
    dc = date_cols[0]
    d = df[[dc]].dropna().copy()
    if len(d) < 5:
        return None
    d["_week"] = d[dc].dt.to_period("W").dt.start_time
    g = d.groupby("_week").size().reset_index(name="Rows")
    fig = px.line(g, x="_week", y="Rows", markers=True)
    fig.update_traces(line=dict(color=PALETTE[0], width=2.2))
    fig = _fig_layout(fig, f"Row volume per week — {dc}")
    fig.update_xaxes(title_text=dc)
    fig.update_yaxes(title_text="Count of rows")
    return _to_fig_dict(fig)


def detect_outliers(df: pd.DataFrame) -> pd.DataFrame:
    numeric = df.select_dtypes(include=[np.number])
    if numeric.empty:
        return pd.DataFrame()
    mask = pd.Series(False, index=df.index)
    reasons = pd.Series("", index=df.index, dtype=object)
    for col in numeric.columns:
        q1, q3 = numeric[col].quantile(0.25), numeric[col].quantile(0.75)
        iqr = q3 - q1
        lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        col_mask = (numeric[col] < lo) | (numeric[col] > hi)
        reasons.loc[col_mask] = reasons.loc[col_mask].astype(str) + f"{col} "
        mask |= col_mask
    out = df[mask].copy()
    if not out.empty:
        out["_outlier_cols"] = reasons[mask].str.strip()
    return out.head(100)


def compute_grounded_stats(df: pd.DataFrame) -> dict:
    """Compute authoritative summary stats the AI narrative MUST reference.

    The AI is given these numbers so its prose can be checked against the
    data rather than fabricating totals. All calculations skip NaN (pandas
    default) — we never fill nulls upstream.
    """
    id_cols = set(_detect_id_columns(df))
    numeric = [c for c in df.select_dtypes(include=[np.number]).columns if c not in id_cols]
    date_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64"]).columns.tolist()

    stats_out: dict = {
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "id_columns": list(id_cols),
        "numeric_columns": numeric,
        "date_columns": date_cols,
        "date_range": None,
        "totals": {},          # sum/mean/median per measure column
        "top_categorical": {}, # {col: [(value, count), ...]}
        "group_tests": [],     # ANOVA results per category x primary metric
    }

    if date_cols:
        dc = date_cols[0]
        d = df[dc].dropna()
        if len(d):
            stats_out["date_range"] = {
                "column": dc,
                "start": str(d.min().date()),
                "end": str(d.max().date()),
                "span_days": int((d.max() - d.min()).days),
            }

    for c in numeric:
        col = df[c].dropna()
        if col.empty:
            continue
        stats_out["totals"][c] = {
            "n": int(col.size),
            "sum": float(col.sum()),
            "mean": float(col.mean()),
            "median": float(col.median()),
            "std": float(col.std(ddof=1)) if col.size > 1 else 0.0,
            "min": float(col.min()),
            "max": float(col.max()),
            "negative": int((col < 0).sum()),
            "zero": int((col == 0).sum()),
        }

    # Top categorical values and ANOVA across them for the primary metric.
    primary = next((c for c in numeric if _is_amount_col(c)),
                   next((c for c in numeric if _is_count_col(c)),
                        numeric[0] if numeric else None))
    for c in df.columns:
        if c in id_cols or not (df[c].dtype == object or pd.api.types.is_bool_dtype(df[c])):
            continue
        nn = df[c].dropna()
        if not (2 <= nn.nunique() <= 30):
            continue
        vc = nn.value_counts().head(10)
        stats_out["top_categorical"][c] = [[str(k), int(v)] for k, v in vc.items()]

        if primary and primary in df.columns:
            groups = [
                df.loc[df[c] == g, primary].dropna().values
                for g in nn.unique() if len(df.loc[df[c] == g, primary].dropna()) >= 3
            ]
            if len(groups) >= 2:
                try:
                    f_stat, p_val = stats.f_oneway(*groups)
                    if np.isfinite(p_val):
                        stats_out["group_tests"].append({
                            "category": c,
                            "metric": primary,
                            "n_groups": len(groups),
                            "f_stat": round(float(f_stat), 4),
                            "p_value": round(float(p_val), 5),
                            "significant_alpha_05": bool(p_val < 0.05),
                        })
                except Exception:
                    pass
    return stats_out


def run_ab_significance(df: pd.DataFrame) -> dict | None:
    """Try to detect variant + metric columns and run a significance test."""
    cols_lower = {c.lower(): c for c in df.columns}
    variant_col = next((cols_lower[k] for k in cols_lower
                        if "variant" in k or "group" in k or k in ("a_b", "ab", "test")), None)
    if not variant_col:
        for c in df.columns:
            if df[c].nunique() == 2 and df[c].dtype == object:
                variant_col = c
                break
    if not variant_col:
        return None

    groups = df[variant_col].dropna().unique()
    if len(groups) != 2:
        return None

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if not numeric_cols:
        return None
    metric = numeric_cols[0]

    a = df[df[variant_col] == groups[0]][metric].dropna()
    b = df[df[variant_col] == groups[1]][metric].dropna()
    if len(a) < 10 or len(b) < 10:
        return {"summary": "Sample too small for reliable A/B test.", "significant": False}

    t_stat, p_val = stats.ttest_ind(a, b, equal_var=False)
    mean_a, mean_b = a.mean(), b.mean()
    lift = ((mean_b - mean_a) / mean_a * 100) if mean_a else 0
    significant = p_val < 0.05
    winner = groups[1] if mean_b > mean_a else groups[0]

    return {
        "variant_col": variant_col,
        "metric": metric,
        "group_a": str(groups[0]), "mean_a": round(float(mean_a), 4), "n_a": int(len(a)),
        "group_b": str(groups[1]), "mean_b": round(float(mean_b), 4), "n_b": int(len(b)),
        "t_stat": round(float(t_stat), 4),
        "p_value": round(float(p_val), 5),
        "lift_pct": round(float(lift), 2),
        "significant": bool(significant),
        "winner": str(winner),
        "summary": (
            f"Variant '{winner}' has a mean {metric} of "
            f"{(mean_b if winner == groups[1] else mean_a):.3f} vs "
            f"{(mean_a if winner == groups[1] else mean_b):.3f} — "
            f"a {abs(lift):.1f}% {'lift' if lift > 0 else 'decline'}. "
            f"p = {p_val:.4f} — {'statistically significant' if significant else 'NOT statistically significant'} at α=0.05."
        ),
    }


# -----------------------------------------------------------------------------
# SQL generation
# -----------------------------------------------------------------------------
SNOWFLAKE_TYPE_MAP = {
    "int64": "NUMBER", "int32": "NUMBER", "float64": "FLOAT", "float32": "FLOAT",
    "bool": "BOOLEAN",
    "datetime64[ns]": "TIMESTAMP_NTZ",
    "object": "VARCHAR",
}


def sql_create_table(df: pd.DataFrame, table_name: str = "dataset") -> str:
    cols = []
    for c in df.columns:
        dtype = str(df[c].dtype)
        sql_type = SNOWFLAKE_TYPE_MAP.get(dtype, "VARCHAR")
        clean_name = '"' + str(c).replace('"', '') + '"'
        cols.append(f"  {clean_name} {sql_type}")
    return f"CREATE OR REPLACE TABLE dataset (\n" + ",\n".join(cols) + "\n);"


# -----------------------------------------------------------------------------
# Routes
# -----------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    return jsonify({"ok": True, "ts": datetime.utcnow().isoformat()})


def detect_provider(api_key: str) -> str | None:
    """Infer provider from API key format.

    - Anthropic keys: 'sk-ant-...'
    - Google Gemini keys: 'AIza...'  (Google API key pattern)
    - OpenAI keys: 'sk-...' or 'sk-proj-...'
    """
    if not api_key:
        return None
    k = api_key.strip()
    if k.startswith("sk-ant-"):
        return "anthropic"
    if k.startswith("AIza"):
        return "gemini"
    if k.startswith("sk-"):
        return "openai"
    return None


@app.route("/api/config", methods=["POST"])
def api_config():
    data = request.get_json(silent=True) or {}
    api_key = (data.get("api_key") or "").strip()
    if not api_key or len(api_key) < 10:
        return jsonify({"error": "API key missing or too short"}), 400

    provider = detect_provider(api_key)
    if not provider:
        return jsonify({
            "error": "Unrecognized API key format. Expected an OpenAI (sk-...), Anthropic (sk-ant-...), or Google Gemini (AIza...) key.",
        }), 400

    state = get_state()
    state["provider"] = provider
    state["api_key"] = api_key

    labels = {
        "openai": "OpenAI · gpt-4o",
        "anthropic": "Anthropic · claude-sonnet-4",
        "gemini": "Google · gemini-1.5-pro",
    }
    return jsonify({"ok": True, "provider": provider, "label": labels[provider]})


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    filename = secure_filename(f.filename or "dataset.csv")

    try:
        raw = f.read()
        # Try common encodings
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return jsonify({"error": "Could not decode CSV. Try UTF-8 encoding."}), 400

        if df.empty:
            return jsonify({"error": "CSV is empty"}), 400

        state = get_state()
        state["original_df"] = df.copy()
        state["filename"] = filename
        return jsonify({
            "ok": True,
            "filename": filename,
            "rows": int(len(df)),
            "cols": int(len(df.columns)),
            "columns": list(df.columns),
        })
    except pd.errors.ParserError as e:
        return jsonify({"error": f"Malformed CSV: {str(e)[:200]}"}), 400
    except Exception as e:
        log.error(f"Upload failed: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Upload failed: {str(e)[:200]}"}), 500


@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    state = get_state()
    body = request.get_json(silent=True) or {}

    # Accept api_key from request body (preferred, survives server restarts)
    # or fall back to server session state.
    api_key = (body.get("api_key") or state.get("api_key") or "").strip()
    provider = detect_provider(api_key) if api_key else state.get("provider")

    if "original_df" not in state:
        return jsonify({"error": "No dataset found on the server. Please re-upload your CSV."}), 400
    if not api_key or not provider:
        return jsonify({"error": "API key missing or unrecognized. Please re-enter your key and click Connect."}), 400

    # Cache for subsequent calls (chat, etc.)
    state["api_key"] = api_key
    state["provider"] = provider

    mode = body.get("mode", "general")
    custom = body.get("custom", "")
    benchmarks = body.get("benchmarks", [])  # list of {metric, value}
    archetype_override = (body.get("archetype_override") or "").strip().lower()

    try:
        # 1. Clean
        df, clean_summary = clean_dataframe(state["original_df"].copy())
        state["cleaned_df"] = df
        state["clean_summary"] = clean_summary
        # New dataset — drop any cached RAG evidence index from a prior run.
        state.pop("evidence_index", None)

        # 2. Profile
        profile = profile_dataframe(df)

        # 3. Detect B2C archetype + run the matching senior-analyst playbook.
        # This is the specialization layer: instead of generic stats, we
        # apply the right analytical playbook (RFM/cohorts for orders,
        # ROAS/CAC for marketing, etc.) for the detected dataset type.
        archetype = arch_mod.detect_archetype(df)
        # User override: if the detector got it wrong, the UI's dropdown
        # forces a different archetype. We keep the role-column mapping
        # since those rules are independent of the archetype label.
        if archetype_override and archetype_override in playbook_mod.PLAYBOOKS:
            archetype = arch_mod.ArchetypeMatch(
                name=archetype_override,
                confidence=1.0,  # user-asserted
                signals=[f"user override (was {archetype.name})"] + archetype.signals,
                role_columns=archetype.role_columns,
            )
        playbook = playbook_mod.run_playbook(archetype.name, df, archetype.role_columns)
        state["archetype"] = archetype
        state["playbook"] = playbook

        # 4. Compute authoritative summary stats BEFORE the AI runs. These
        # numbers are the ground truth the narrative must reference.
        grounded = compute_grounded_stats(df)

        # 5. AI narrative — given the archetype + playbook output so it
        # speaks the right analyst vocabulary and cites real numbers.
        ai = analyze(provider, api_key, profile, mode, custom, benchmarks,
                     grounded=grounded, clean_summary=clean_summary,
                     archetype=archetype, playbook=playbook)

        # 5. Build charts deterministically in code. Archetype-aware charts
        # (cohort heatmap, channel matrix, funnel, etc.) come first, generic
        # ones (top-N, scatter) fill in behind them.
        arch_charts = playbook_mod.build_archetype_charts(
            archetype.name, df, archetype.role_columns, playbook)
        generic_charts = build_auto_charts(df)
        # Don't duplicate concepts the archetype chart already covered well
        # (e.g. don't show two "channel" bars). Cheap heuristic by title prefix.
        arch_titles_lower = {c.get("title", "").lower() for c in arch_charts}
        generic_charts = [c for c in generic_charts
                          if c.get("title", "").lower() not in arch_titles_lower]
        charts = arch_charts + generic_charts

        # 6. Auto features
        corr = correlation_heatmap(df)
        ts = time_series_trend(df)
        outliers_df = detect_outliers(df)
        ab = run_ab_significance(df) if mode == "abtest" or mode == "general" else None

        # 6. SQL
        create_sql = sql_create_table(df)
        sql_queries = [{"title": "CREATE TABLE", "sql": create_sql}]
        sql_queries.extend(ai.get("sql_queries", []))

        # 7. Benchmark overlays — simple attachment to response, JS renders reference lines
        state["last_analysis"] = {
            "mode": mode,
            "custom": custom,
            "benchmarks": benchmarks,
            "ai": ai,
            "charts": charts,
            "correlation": corr,
            "timeseries": ts,
            "ab_test": ab,
            "sql_queries": sql_queries,
            "outliers_count": len(outliers_df),
        }

        # 8. Response payload
        return jsonify({
            "ok": True,
            "filename": state.get("filename"),
            "rows": int(len(df)),
            "cols": int(len(df.columns)),
            "mode": TEMPLATES.get(mode, TEMPLATES["general"])["name"],
            "archetype": {
                "name": archetype.name,
                "confidence": archetype.confidence,
                "description": arch_mod.archetype_description(archetype.name),
                "signals": archetype.signals,
                "role_columns": archetype.role_columns,
            },
            "playbook": playbook,
            "clean_summary": clean_summary,
            "profile": {
                "columns": profile["columns"],
                "dtypes": profile["dtypes"],
                "null_counts": profile["null_counts"],
                "nunique": profile["nunique"],
            },
            "executive_summary": ai.get("executive_summary", ""),
            "kpi_cards": ai.get("kpi_cards", []),
            "data_quality_notes": ai.get("data_quality_notes", []),
            "followup_questions": ai.get("followup_questions", []),
            "charts": charts,
            "correlation": corr,
            "timeseries": ts,
            "ab_test": ab,
            "outliers": {
                "count": int(len(outliers_df)),
                "rows": outliers_df.head(50).astype(str).to_dict(orient="records"),
            },
            "sql_queries": sql_queries,
            "benchmarks": benchmarks,
            "grounded_stats": grounded,
            "preview": df.head(200).astype(str).to_dict(orient="records"),
        })
    except Exception as e:
        log.error(f"Analysis failed: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Analysis failed: {str(e)[:300]}"}), 500


@app.route("/api/column/<name>", methods=["GET"])
def api_column(name):
    state = get_state()
    df = state.get("cleaned_df")
    if df is None or name not in df.columns:
        return jsonify({"error": "Column not found"}), 404

    s = df[name]
    info = {
        "name": name,
        "dtype": str(s.dtype),
        "null_pct": round(float(s.isna().mean() * 100), 2),
        "unique": int(s.nunique()),
        "total": int(len(s)),
    }
    if pd.api.types.is_numeric_dtype(s):
        info.update({
            "min": float(s.min()) if s.notna().any() else None,
            "max": float(s.max()) if s.notna().any() else None,
            "mean": round(float(s.mean()), 4) if s.notna().any() else None,
            "median": round(float(s.median()), 4) if s.notna().any() else None,
        })
        fig = px.histogram(s.dropna(), nbins=30)
    else:
        top = s.value_counts().head(15).reset_index()
        top.columns = [name, "count"]
        fig = px.bar(top, x=name, y="count")

    fig = _fig_layout(fig, f"{name} distribution")
    info["figure"] = json.loads(json.dumps(fig.to_dict(), cls=PlotlyJSONEncoder))
    return jsonify(info)


@app.route("/api/chat", methods=["POST"])
def api_chat():
    state = get_state()
    df = state.get("cleaned_df")
    if df is None:
        return jsonify({"error": "No dataset analyzed yet"}), 400

    body = request.get_json(silent=True) or {}
    api_key = (body.get("api_key") or state.get("api_key") or "").strip()
    provider = detect_provider(api_key) if api_key else state.get("provider")
    if not api_key or not provider:
        return jsonify({"error": "API key missing. Please re-enter your key."}), 400

    question = (body.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Empty question"}), 400

    profile = profile_dataframe(df)
    prompt = f"""You are a data analyst. Answer the user's question in 2-4 sentences of plain English based on this dataset summary.

DATASET SUMMARY:
{json.dumps(profile, default=str)[:6000]}

USER QUESTION: {question}

Return a JSON object: {{"answer": "your plain-English answer"}}"""

    try:
        caller = {
            "openai": call_openai,
            "anthropic": call_anthropic,
            "gemini": call_gemini,
        }[provider]
        raw = caller(api_key, prompt, strict=False)
        try:
            parsed = _safe_json_extract(raw)
            return jsonify({"ok": True, "answer": parsed.get("answer", raw)})
        except Exception:
            return jsonify({"ok": True, "answer": raw.strip()})
    except Exception as e:
        log.error(f"Chat failed: {e}")
        return jsonify({"error": f"Chat failed: {str(e)[:200]}"}), 500


# -----------------------------------------------------------------------------
# ReturnLens — RAG over customer voice (reviews / tickets / return reasons)
# -----------------------------------------------------------------------------
@app.route("/api/returns_stats", methods=["GET"])
def api_returns_stats():
    """Compute return-rate stats from the already-analyzed dataset.

    Looks for return-related columns (is_return flag, return_reason, etc.)
    and rolls them up overall and per-SKU.
    """
    state = get_state()
    df = state.get("cleaned_df")
    if df is None:
        return jsonify({"error": "No dataset analyzed yet. Run analysis first."}), 400

    # Best-guess SKU column from the dataset.
    sku_col = None
    for c in df.columns:
        cl = str(c).lower()
        if "sku" in cl or "product_id" in cl or cl == "product" or "item_id" in cl:
            sku_col = c
            break

    stats_out = rag.compute_returns_stats(df, sku_col=sku_col)
    stats_out["sku_column"] = sku_col
    return jsonify({"ok": True, "stats": stats_out})


@app.route("/api/ingest_corpus", methods=["POST"])
def api_ingest_corpus():
    """Ingest a customer-voice corpus (reviews / tickets / return reasons).

    Body (multipart): file=<CSV with at least a text column>
                       openai_key=<key for embeddings>
                       source_type=<review|ticket|return_reason>
                       mapping=<optional JSON of column overrides>
    OR
    Body (JSON):       {openai_key, rows: [...], mapping?, source_type?}

    Builds an in-memory embedded index keyed by the user's session. The
    index is consulted later by /api/diagnose.
    """
    state = get_state()
    source_type = (request.form.get("source_type")
                   or (request.get_json(silent=True) or {}).get("source_type")
                   or "review")
    openai_key = (request.form.get("openai_key")
                  or (request.get_json(silent=True) or {}).get("openai_key")
                  or state.get("openai_key") or "").strip()

    if not openai_key:
        return jsonify({"error": "OpenAI API key required for embeddings (text-embedding-3-small)."}), 400
    if not openai_key.startswith("sk-"):
        return jsonify({"error": "That doesn't look like an OpenAI key. ReturnLens uses OpenAI embeddings; please supply an OpenAI key starting with sk-."}), 400

    # Load the CSV.
    try:
        if "file" in request.files:
            f = request.files["file"]
            raw = f.read()
            df = None
            for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(io.BytesIO(raw), encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                return jsonify({"error": "Could not decode corpus CSV."}), 400
        else:
            body = request.get_json(silent=True) or {}
            rows = body.get("rows") or []
            if not rows:
                return jsonify({"error": "No corpus rows provided."}), 400
            df = pd.DataFrame(rows)

        if df.empty:
            return jsonify({"error": "Corpus CSV is empty."}), 400

        # Mapping override from client, or auto-detect.
        mapping = None
        if "mapping" in request.form:
            try:
                mapping = json.loads(request.form["mapping"])
            except json.JSONDecodeError:
                mapping = None
        if not mapping:
            body = request.get_json(silent=True) or {}
            mapping = body.get("mapping")
        if not mapping:
            mapping = rag.detect_corpus_columns(df)

        chunks = rag.rows_to_chunks(df, mapping, default_source=source_type)
        if not chunks:
            return jsonify({
                "error": "No usable text rows found. Make sure your corpus CSV "
                         "has a column with review/comment/ticket text.",
                "detected_mapping": mapping,
            }), 400

        # Cap to keep embedding cost / latency bounded on a free Render tier.
        MAX_CHUNKS = 2000
        if len(chunks) > MAX_CHUNKS:
            chunks = chunks[:MAX_CHUNKS]

        index = rag.build_index(chunks, openai_api_key=openai_key)

        # Merge with any prior corpus the user uploaded in this session,
        # rather than replacing. Useful for combining reviews + tickets.
        prior = state.get("rag_index")
        if prior and len(prior.get("chunks", [])):
            merged_chunks = list(prior["chunks"]) + chunks
            merged_emb = np.vstack([prior["embeddings"], index["embeddings"]])
            index = {
                "chunks": merged_chunks,
                "embeddings": merged_emb,
                "built_at": datetime.utcnow().isoformat(),
            }

        state["rag_index"] = index
        state["openai_key"] = openai_key
        state["corpus_mapping"] = mapping

        # Useful summary stats for the UI panel.
        n = len(index["chunks"])
        sku_counts: dict = {}
        rating_counts: dict = {}
        for c in index["chunks"]:
            if c.get("sku"):
                sku_counts[c["sku"]] = sku_counts.get(c["sku"], 0) + 1
            if c.get("rating") is not None:
                key = str(int(c["rating"]))
                rating_counts[key] = rating_counts.get(key, 0) + 1

        return jsonify({
            "ok": True,
            "indexed": n,
            "source_type": source_type,
            "mapping": mapping,
            "top_skus": sorted(sku_counts.items(), key=lambda kv: -kv[1])[:10],
            "rating_distribution": rating_counts,
        })
    except Exception as e:
        log.error(f"Corpus ingestion failed: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Ingestion failed: {str(e)[:200]}"}), 500


@app.route("/api/diagnose", methods=["POST"])
def api_diagnose():
    """Grounded diagnostic Q&A.

    Combines (a) numerical return-rate stats from the CSV and (b) retrieved
    customer-voice chunks. The LLM is constrained to cite both sources and
    forbidden from inventing quotes.

    Body: {api_key, question, sku?, min_rating?, max_rating?, source_type?, k?}
    """
    state = get_state()
    body = request.get_json(silent=True) or {}

    api_key = (body.get("api_key") or state.get("api_key") or "").strip()
    openai_key = (body.get("openai_key") or state.get("openai_key") or "").strip()
    provider = detect_provider(api_key) if api_key else state.get("provider")

    if state.get("cleaned_df") is None:
        return jsonify({"error": "No analyzed dataset. Run analysis first."}), 400
    if not state.get("rag_index"):
        return jsonify({"error": "No customer-voice corpus uploaded yet. Upload reviews or tickets first."}), 400
    if not api_key or not provider:
        return jsonify({"error": "Narrative LLM key missing (Claude / OpenAI / Gemini)."}), 400
    if not openai_key:
        return jsonify({"error": "OpenAI key required for embedding the question."}), 400

    question = (body.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Empty question."}), 400

    sku = (body.get("sku") or "").strip() or None
    min_rating = body.get("min_rating")
    max_rating = body.get("max_rating")
    source_type = body.get("source_type")
    k = int(body.get("k") or 6)

    # Build filter dict for metadata pre-filter.
    filters: dict = {}
    if sku:
        filters["sku"] = ("eq", sku)
    if source_type:
        filters["source_type"] = ("eq", source_type)
    if min_rating is not None:
        filters["rating"] = ("gte", min_rating)
    if max_rating is not None:
        # Both gte and lte on the same field — apply max as a second pass.
        existing = filters.get("rating")
        if existing and existing[0] == "gte":
            # combine via a tuple-of-tuples we'll handle below
            filters.pop("rating")
            filters["__rating_gte"] = existing[1]
            filters["__rating_lte"] = max_rating
        else:
            filters["rating"] = ("lte", max_rating)

    # The custom op encoding above isn't supported by retrieve directly;
    # so just split into two passes if both bounds are present.
    def _post_rating_filter(rows):
        lo = filters.pop("__rating_gte", None)
        hi = filters.pop("__rating_lte", None)
        if lo is None and hi is None:
            return rows
        out = []
        for r in rows:
            v = r.get("rating")
            if v is None:
                continue
            if lo is not None and v < lo:
                continue
            if hi is not None and v > hi:
                continue
            out.append(r)
        return out

    try:
        retrieved = rag.retrieve(
            state["rag_index"], question, openai_api_key=openai_key,
            k=max(k, 12), filters={k_: v for k_, v in filters.items() if not k_.startswith("__")},
        )
        retrieved = _post_rating_filter(retrieved)[:k]
    except Exception as e:
        log.error(f"Retrieval failed: {e}")
        return jsonify({"error": f"Retrieval failed: {str(e)[:200]}"}), 500

    # CSV-grounded numerical context: return-rate for the queried SKU if any.
    df = state["cleaned_df"]
    sku_col = None
    for c in df.columns:
        cl = str(c).lower()
        if "sku" in cl or "product_id" in cl or "item_id" in cl:
            sku_col = c
            break
    csv_stats = rag.compute_returns_stats(df, sku_col=sku_col)
    if sku and sku_col and csv_stats.get("per_sku"):
        per_sku_row = next(
            (r for r in csv_stats["per_sku"] if str(r["sku"]).lower() == sku.lower()),
            None,
        )
        csv_stats["focus_sku"] = per_sku_row

    # Build the grounded prompt.
    quotes_block = "\n".join([
        f"[Q{i+1}] sku={c.get('sku','?')} rating={c.get('rating','?')} "
        f"date={c.get('date','?')} source={c.get('source_type','?')}: "
        f"\"{c['text'][:500]}\""
        for i, c in enumerate(retrieved)
    ]) or "(no relevant customer quotes retrieved)"

    prompt = f"""You are a returns and customer-experience analyst for a DTC brand.
Answer the user's question by combining (a) the CSV-grounded numerical stats below
and (b) the retrieved customer quotes. STRICT RULES:

1. Every numerical claim must come from the CSV STATS block. Do NOT invent numbers.
2. Every claim about what customers say or feel must be backed by a quote from the
   RETRIEVED QUOTES block, cited inline as [Q1], [Q2], etc.
3. Do NOT fabricate quotes. If the retrieved quotes don't support a confident
   answer, say so plainly.
4. End your answer with a single concrete "Recommended fix" the brand could ship
   this week.

CSV STATS:
{json.dumps(csv_stats, default=str)[:4000]}

RETRIEVED QUOTES ({len(retrieved)} of corpus, filtered by {json.dumps({k_: v for k_, v in filters.items() if not k_.startswith('__')}) or 'no filters'}):
{quotes_block}

USER QUESTION: {question}

Return a JSON object: {{
  "headline": "one-sentence diagnosis",
  "stats_summary": "1-2 sentences citing CSV stats",
  "voice_summary": "2-4 sentences citing [Q#] quotes",
  "recommended_fix": "one concrete shippable action",
  "confidence": "high|medium|low"
}}"""

    try:
        caller = {"openai": call_openai, "anthropic": call_anthropic,
                  "gemini": call_gemini}[provider]
        raw = caller(api_key, prompt, strict=False)
        try:
            parsed = _safe_json_extract(raw)
        except Exception:
            parsed = {"headline": raw[:400], "stats_summary": "",
                      "voice_summary": "", "recommended_fix": "",
                      "confidence": "low"}
        return jsonify({
            "ok": True,
            "answer": parsed,
            "citations": [
                {"id": f"Q{i+1}", **c} for i, c in enumerate(retrieved)
            ],
            "csv_stats": csv_stats,
        })
    except Exception as e:
        log.error(f"Diagnose failed: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Diagnose failed: {str(e)[:200]}"}), 500


# -----------------------------------------------------------------------------
# Synthetic Research — Phase 2: persona-grounded synthetic market surveys
# -----------------------------------------------------------------------------
def _brand_from_filename(fn: str | None) -> str:
    """Derive a readable brand name from the uploaded filename."""
    if not fn:
        return "the brand"
    base = fn.rsplit(".", 1)[0].replace("_", " ").replace("-", " ")
    stop = {"sales", "data", "cleaned", "uncleaned", "dataset", "export",
            "final", "csv", "orders", "customers", "2", "v2", "copy"}
    words = [w for w in base.split() if w.lower() not in stop]
    return " ".join(words[:2]).title() if words else "the brand"


def _load_calibration_profile() -> dict | None:
    """Read calibration_profile.json if calibration.py has been run."""
    try:
        with open("calibration_profile.json") as f:
            return json.load(f)
    except Exception:
        return None


def _attach_calibration(result: dict, study_type: str) -> dict:
    """Attach measured backtest accuracy to a synthetic result and, when the
    calibration backtests graded this study type as less trustworthy than
    the result claims, cap the confidence honestly."""
    if not result.get("ok"):
        return result
    prof = _load_calibration_profile()
    if not prof:
        result["calibration"] = {
            "status": "uncalibrated",
            "message": ("Engine not yet validated against real datasets. "
                        "Run calibration.py to backtest it."),
        }
        return result

    # Map each study type to the backtest family that validates it.
    trust_key = {"pricing": "pricing", "comparison": "comparison",
                 "concept": "comparison", "conjoint": "conjoint",
                 "van_westendorp": "pricing"}.get(study_type, "conjoint")
    trust = (prof.get("trust") or {}).get(trust_key)
    result["calibration"] = {
        "status": "calibrated",
        "validated_as": trust_key,
        "trust": trust,
        "conjoint": prof.get("conjoint"),
        "comparison": prof.get("comparison"),
        "notes": prof.get("notes", []),
        "generated_at": prof.get("generated_at"),
    }
    # Honesty cap: never let a result claim more confidence than the
    # backtests support.
    rank = {"low": 0, "medium": 1, "high": 2}
    if trust and rank.get(trust, 1) < rank.get(result.get("confidence", "medium"), 1):
        result["confidence"] = trust
        result.setdefault("caveats", []).append(
            f"Confidence capped at '{trust}' — that is the trust grade the "
            f"calibration backtests measured for {trust_key} studies.")
    return result


@app.route("/api/synthetic_research", methods=["POST"])
def api_synthetic_research():
    """Run a persona-grounded synthetic survey on the analyzed dataset.

    Requires a prior /api/analyze (cleaned_df + archetype must be in
    session). Builds digital-twin segment personas, runs the requested
    study type, returns aggregated results with caveats + confidence.
    """
    state = get_state()
    body = request.get_json(silent=True) or {}

    api_key = (body.get("api_key") or state.get("api_key") or "").strip()
    provider = detect_provider(api_key) if api_key else state.get("provider")
    df = state.get("cleaned_df")
    archetype = state.get("archetype")

    if df is None or archetype is None:
        return jsonify({"error": "Run an analysis first — synthetic research "
                                 "needs the cleaned dataset and detected "
                                 "archetype."}), 400
    if not api_key or not provider:
        return jsonify({"error": "AI key missing or unrecognized."}), 400

    study_type = (body.get("study_type") or "").strip().lower()
    config = body.get("config") or {}
    if study_type not in ("pricing", "concept", "comparison",
                          "conjoint", "van_westendorp"):
        return jsonify({"error": f"Unknown study type: {study_type}"}), 400

    roles = getattr(archetype, "role_columns", {}) or {}
    playbook = state.get("playbook")

    try:
        prof = personas_mod.build_segment_profiles(df, roles, playbook)
        profiles = prof.get("profiles", [])
        if len(profiles) < 2:
            return jsonify({"error": prof.get("note", "Could not build at "
                            "least 2 customer segments to survey. Synthetic "
                            "research needs a segmentable dimension.")}), 400

        demand = personas_mod.fit_demand_curve(df, roles)
        caller = {"openai": call_openai, "anthropic": call_anthropic,
                  "gemini": call_gemini}[provider]
        brand = _brand_from_filename(state.get("filename"))

        # RAG evidence layer: index the real dataset rows so each persona's
        # prompt can be grounded in retrieved real records. Built once and
        # cached in the session. Uses OpenAI embeddings when a key is
        # available; otherwise keyword retrieval.
        openai_key = (body.get("openai_key") or state.get("openai_key")
                      or (api_key if provider == "openai" else None))
        ev_index = state.get("evidence_index")
        if ev_index is None:
            seg_col, seg_lab = personas_mod.segment_labels(df, roles)
            ev_index = evidence_mod.build_evidence_index(
                df, roles, seg_col, seg_lab, openai_key=openai_key)
            state["evidence_index"] = ev_index

        result = sr_mod.run_study(study_type, config, profiles, demand,
                                  caller, api_key, brand,
                                  evidence_index=ev_index, openai_key=openai_key)
        result = _attach_calibration(result, study_type)
        result["rag"] = {
            "grounded": bool(ev_index and ev_index.get("chunks")),
            "n_records_indexed": ev_index.get("n_records", 0) if ev_index else 0,
            "retrieval": ("semantic embeddings" if ev_index
                          and ev_index.get("embedded") else "keyword"),
        }
        result["segmentation_column"] = prof.get("segmentation_column")
        result["panel"] = [
            {"segment": p["name"], "weight": p.get("weight"),
             "n_rows": p.get("n_rows"), "description": p.get("description")}
            for p in profiles
        ]
        if demand.get("usable"):
            result["demand_curve"] = {
                "elasticity": demand.get("elasticity"),
                "r_squared": demand.get("r_squared"),
                "price_band": [demand.get("price_min"), demand.get("price_max")],
            }
        status = 200 if result.get("ok") else 400
        return jsonify(result), status
    except Exception as e:
        log.error(f"Synthetic research failed: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Synthetic research failed: {str(e)[:200]}"}), 500


# -----------------------------------------------------------------------------
# Exports
# -----------------------------------------------------------------------------
@app.route("/api/export/excel", methods=["POST", "GET"])
def export_excel():
    """Prefers POST body (client-supplied payload, survives server restarts);
    falls back to server session state if POST body is empty."""
    state = get_state()
    body = request.get_json(silent=True) or {}

    df = None
    if body.get("rows") and body.get("columns"):
        # Client-supplied cleaned dataset (list of dicts) + column order
        df = pd.DataFrame(body["rows"], columns=body["columns"])
    else:
        df = state.get("cleaned_df")

    last = body.get("last_analysis") or state.get("last_analysis", {})
    clean = body.get("clean_summary") or state.get("clean_summary", {})
    filename = body.get("filename") or state.get("filename") or "dataset.csv"

    if df is None:
        return jsonify({"error": "No analysis data provided. Re-run the analysis first."}), 400

    wb = Workbook()

    # Summary tab
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="15803D")

    ws["A1"] = "AI Analytics Report"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A2"] = f"Dataset: {filename}"
    ws["A3"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

    ws["A5"] = "Executive Summary"
    ws["A5"].font = header_font
    ws["A5"].fill = header_fill
    ws["A6"] = last.get("ai", {}).get("executive_summary", "")
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[6].height = 80
    ws.column_dimensions["A"].width = 100

    row = 8
    ws.cell(row=row, column=1, value="KPI Cards").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="Label").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Detail").font = Font(bold=True)
    row += 1
    for kpi in last.get("ai", {}).get("kpi_cards", []):
        ws.cell(row=row, column=1, value=str(kpi.get("label", "")))
        ws.cell(row=row, column=2, value=str(kpi.get("value", "")))
        ws.cell(row=row, column=3, value=str(kpi.get("subtext", "")))
        row += 1

    # Cleaned Data tab
    ws2 = wb.create_sheet("Cleaned Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data Quality tab
    ws3 = wb.create_sheet("Data Quality")
    ws3["A1"] = "Cleaning Report"
    ws3["A1"].font = header_font
    ws3["A1"].fill = header_fill
    r = 3
    for key, label in [
        ("duplicates_removed", "Duplicate rows removed"),
        ("duplicate_key", "Deduplication key"),
        ("id_columns", "Primary-key-like columns"),
        ("whitespace_columns_fixed", "Columns with whitespace trimmed"),
        ("types_inferred", "Columns with type inference applied"),
        ("category_merges", "Categorical labels merged (typo/case)"),
        ("high_null_columns", "High-null columns (>=30% null, KEPT)"),
        ("negative_in_positive_cols", "Negative values in expected-positive columns"),
        ("zero_in_positive_cols", "Zero values in expected-positive columns"),
        ("suspect_negatives_nulled", "Negative values nullified (set to NaN)"),
        ("suspect_zeros_nulled", "Zero values nullified (set to NaN)"),
        ("invalid_rates_nulled", "Out-of-range rate/discount values nullified"),
        ("revenue_reconstructed", "Revenue rows reconstructed from components"),
        ("rows_with_any_null", "Rows with at least one null"),
    ]:
        val = clean.get(key)
        ws3.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r, column=2, value=json.dumps(val) if val else "None")
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"analytics_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/export/pdf", methods=["POST", "GET"])
def export_pdf():
    """Prefers POST body (client payload, survives server restarts);
    falls back to server session state if body is empty."""
    state = get_state()
    body = request.get_json(silent=True) or {}

    df = None
    if body.get("rows") and body.get("columns"):
        df = pd.DataFrame(body["rows"], columns=body["columns"])
    else:
        df = state.get("cleaned_df")

    last = body.get("last_analysis") or state.get("last_analysis")
    clean_body = body.get("clean_summary")
    filename_override = body.get("filename")

    if df is None or not last:
        return jsonify({"error": "No analysis data provided. Re-run the analysis first."}), 400

    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=LETTER,
            leftMargin=0.7 * inch, rightMargin=0.7 * inch,
            topMargin=0.7 * inch, bottomMargin=0.7 * inch,
            title="AI Analytics Report",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title", parent=styles["Title"], fontSize=28,
            textColor=colors.HexColor(ACCENT), spaceAfter=20,
        )
        h2 = ParagraphStyle(
            "H2", parent=styles["Heading2"], fontSize=16,
            textColor=colors.HexColor(ACCENT), spaceBefore=14, spaceAfter=8,
        )
        body = styles["BodyText"]

        story = []

        # Title page
        story.append(Spacer(1, 1.5 * inch))
        story.append(Paragraph("AI Analytics Report", title_style))
        story.append(Paragraph(
            f"Dataset: <b>{filename_override or state.get('filename', '—')}</b>", body))
        story.append(Paragraph(
            f"Rows: {len(df):,} &nbsp;&nbsp; Columns: {len(df.columns)}", body))
        story.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", body))
        story.append(PageBreak())

        # Executive summary
        story.append(Paragraph("Executive Summary", h2))
        story.append(Paragraph(
            last.get("ai", {}).get("executive_summary", "").replace("\n", "<br/>"),
            body,
        ))

        # KPI cards
        kpis = last.get("ai", {}).get("kpi_cards", [])
        if kpis:
            story.append(Paragraph("Key Metrics", h2))
            data = [["Metric", "Value", "Detail"]]
            for k in kpis:
                data.append([
                    str(k.get("label", "")),
                    str(k.get("value", "")),
                    str(k.get("subtext", "")),
                ])
            t = Table(data, colWidths=[2.0 * inch, 1.6 * inch, 3.2 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(ACCENT)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]))
            story.append(t)

        # Cleaning report
        story.append(Paragraph("Data Cleaning Report", h2))
        clean = clean_body or state.get("clean_summary", {})
        clean_rows = [
            ["Original shape", str(clean.get("original_shape"))],
            ["Cleaned shape", str(clean.get("cleaned_shape"))],
            ["Duplicates removed", f"{clean.get('duplicates_removed', 0)} (key: {clean.get('duplicate_key', '—')})"],
            ["ID columns", ", ".join(clean.get("id_columns", [])) or "—"],
            ["Types inferred", json.dumps(clean.get("types_inferred", {}))[:400]],
            ["Category merges", json.dumps(clean.get("category_merges", {}))[:400]],
            ["High-null columns (kept)", json.dumps(clean.get("high_null_columns", {}))[:400]],
            ["Negative in positive cols", json.dumps(clean.get("negative_in_positive_cols", {}))[:400]],
            ["Zero in positive cols", json.dumps(clean.get("zero_in_positive_cols", {}))[:400]],
            ["Negatives nullified (→ NaN)", json.dumps(clean.get("suspect_negatives_nulled", {}))[:400]],
            ["Zeros nullified (→ NaN)", json.dumps(clean.get("suspect_zeros_nulled", {}))[:400]],
            ["Invalid rates/discounts nullified", json.dumps(clean.get("invalid_rates_nulled", {}))[:400]],
            ["Revenue reconstructed", json.dumps(clean.get("revenue_reconstructed", {}))[:400]],
            ["Rows with any null", str(clean.get("rows_with_any_null", 0))],
        ]
        t = Table(clean_rows, colWidths=[1.8 * inch, 5.0 * inch])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)

        # Charts — embed client-rendered PNGs when available
        story.append(PageBreak())
        story.append(Paragraph("Analyses & Insights", h2))

        def _embed_image(data_url):
            if not data_url or not data_url.startswith("data:image"):
                return None
            try:
                b64 = data_url.split(",", 1)[1]
                raw = base64.b64decode(b64)
                img = RLImage(io.BytesIO(raw), width=6.8 * inch, height=3.83 * inch)
                img.hAlign = "CENTER"
                return img
            except Exception:
                return None

        for c in last.get("charts", []):
            story.append(Paragraph(f"<b>{c.get('title', '')}</b>", body))
            img = _embed_image(c.get("image"))
            if img is not None:
                story.append(Spacer(1, 0.08 * inch))
                story.append(img)
            story.append(Spacer(1, 0.08 * inch))
            story.append(Paragraph(c.get("insight", ""), body))
            story.append(Spacer(1, 0.18 * inch))

        corr_img = _embed_image(last.get("correlation_image"))
        if corr_img is not None:
            story.append(Paragraph("<b>Correlation heatmap</b>", body))
            story.append(corr_img)
            story.append(Spacer(1, 0.18 * inch))
        ts_img = _embed_image(last.get("timeseries_image"))
        if ts_img is not None:
            story.append(Paragraph("<b>Time series overview</b>", body))
            story.append(ts_img)
            story.append(Spacer(1, 0.18 * inch))

        # Quality + follow-ups
        dq = last.get("ai", {}).get("data_quality_notes", [])
        if dq:
            story.append(Paragraph("Data Quality Notes", h2))
            for n in dq:
                story.append(Paragraph(f"• {n}", body))

        fu = last.get("ai", {}).get("followup_questions", [])
        if fu:
            story.append(Paragraph("Recommended Follow-up Questions", h2))
            for q in fu:
                story.append(Paragraph(f"• {q}", body))

        def _page_num(canvas, doc_):
            canvas.saveState()
            canvas.setFont("Helvetica", 9)
            canvas.setFillColor(colors.HexColor("#64748B"))
            canvas.drawRightString(
                LETTER[0] - 0.7 * inch, 0.4 * inch, f"Page {doc_.page}"
            )
            canvas.drawString(0.7 * inch, 0.4 * inch, "AI Analytics Report")
            canvas.restoreState()

        doc.build(story, onFirstPage=_page_num, onLaterPages=_page_num)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True,
            download_name=f"analytics_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf",
            mimetype="application/pdf",
        )
    except Exception as e:
        log.error(f"PDF export failed: {e}\n{traceback.format_exc()}")
        # Fallback HTML print view
        html = "<html><body><h1>AI Analytics Report (fallback)</h1>"
        html += f"<p><b>Dataset:</b> {state.get('filename', '—')}</p>"
        html += f"<h2>Executive Summary</h2><p>{last.get('ai', {}).get('executive_summary', '')}</p>"
        html += "<h2>Insights</h2><ul>"
        for c in last.get("charts", []):
            html += f"<li><b>{c.get('title')}</b>: {c.get('insight')}</li>"
        html += "</ul></body></html>"
        return html, 200, {"Content-Type": "text/html"}


@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large (max 50MB)"}), 413


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
